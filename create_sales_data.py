"""
create_sales_data.py
--------------------
Generates an Excel file called 'supermarket_sales.xlsx'
with 100 rows of fake supermarket sales from around the world.

Run this FIRST to create the Excel file, then run readexcelmain.py to view it.

Requires: pip install openpyxl
"""

import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ── Data pools ──────────────────────────────────────────────────────────────

# Each country maps to a list of cities and a list of supermarket chains
COUNTRY_DATA = {
    "USA":          (["New York", "Los Angeles", "Chicago"],        ["Walmart", "Kroger", "Target"]),
    "UK":           (["London", "Manchester", "Birmingham"],        ["Tesco", "Sainsbury's", "Asda"]),
    "Germany":      (["Berlin", "Munich", "Hamburg"],               ["Aldi", "Lidl", "Rewe"]),
    "France":       (["Paris", "Lyon", "Marseille"],                ["Carrefour", "Leclerc", "Intermarché"]),
    "Japan":        (["Tokyo", "Osaka", "Kyoto"],                   ["Aeon", "Ito-Yokado", "FamilyMart"]),
    "Australia":    (["Sydney", "Melbourne", "Brisbane"],           ["Woolworths", "Coles", "IGA"]),
    "Brazil":       (["Sao Paulo", "Rio de Janeiro", "Brasilia"],   ["Pao de Acucar", "Carrefour", "Extra"]),
    "Canada":       (["Toronto", "Vancouver", "Montreal"],          ["Loblaws", "Sobeys", "Metro"]),
    "India":        (["Mumbai", "Delhi", "Bangalore"],              ["Big Bazaar", "D-Mart", "Reliance Fresh"]),
    "China":        (["Beijing", "Shanghai", "Guangzhou"],          ["RT-Mart", "Hema", "Wumart"]),
    "South Africa": (["Cape Town", "Johannesburg", "Durban"],       ["Shoprite", "Pick n Pay", "Checkers"]),
    "Mexico":       (["Mexico City", "Guadalajara", "Monterrey"],   ["Walmart", "Soriana", "Chedraui"]),
    "UAE":          (["Dubai", "Abu Dhabi", "Sharjah"],             ["Carrefour", "Lulu", "Spinneys"]),
    "Italy":        (["Rome", "Milan", "Naples"],                   ["Esselunga", "Conad", "Coop"]),
    "Russia":       (["Moscow", "St. Petersburg", "Novosibirsk"],   ["Magnit", "X5 Retail", "Lenta"]),
}

# Each category maps to a list of products
CATEGORIES = {
    "Produce":    ["Apples", "Bananas", "Tomatoes", "Spinach", "Carrots"],
    "Dairy":      ["Milk", "Cheese", "Yogurt", "Butter", "Eggs"],
    "Bakery":     ["Bread", "Croissant", "Muffins", "Bagels", "Donuts"],
    "Meat":       ["Chicken", "Beef", "Pork", "Lamb", "Salmon"],
    "Beverages":  ["Orange Juice", "Water", "Coffee", "Tea", "Soda"],
    "Snacks":     ["Chips", "Crackers", "Nuts", "Popcorn", "Cookies"],
    "Frozen":     ["Ice Cream", "Frozen Pizza", "Frozen Veggies", "Fish Fingers", "Waffles"],
    "Household":  ["Detergent", "Toilet Paper", "Shampoo", "Soap", "Tissue"],
}


# ── Helper function ──────────────────────────────────────────────────────────

def random_date_2024():
    """Return a random date within the year 2024."""
    start = date(2024, 1, 1)
    end   = date(2024, 12, 31)
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))


# ── Build 100 rows of data ───────────────────────────────────────────────────

rows = []
countries = list(COUNTRY_DATA.keys())

for order_id in range(1, 101):
    # Pick a random country, then a city and store from that country
    country          = random.choice(countries)
    cities, stores   = COUNTRY_DATA[country]
    city             = random.choice(cities)
    store            = random.choice(stores)

    # Pick a random category and product within it
    category = random.choice(list(CATEGORIES.keys()))
    product  = random.choice(CATEGORIES[category])

    # Generate random sales numbers
    quantity    = random.randint(1, 50)
    unit_price  = round(random.uniform(0.5, 25.0), 2)
    total_sales = round(quantity * unit_price, 2)
    sale_date   = random_date_2024()

    rows.append([order_id, country, city, store, category,
                 product, quantity, unit_price, total_sales, sale_date])


# ── Write to Excel ───────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Supermarket Sales"

# Column headers
headers = [
    "Order_ID", "Country", "City", "Store_Name", "Category",
    "Product", "Quantity", "Unit_Price_USD", "Total_Sales_USD", "Date"
]

# Style: bold white text on a blue background for the header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center")

# Write all data rows (openpyxl appends below the header automatically)
for row_data in rows:
    ws.append(row_data)

# Set reasonable column widths so the Excel is easy to read
column_widths = [10, 15, 18, 22, 12, 16, 10, 16, 17, 12]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Save the file next to this script
wb.save("supermarket_sales.xlsx")
print("Done! 'supermarket_sales.xlsx' created with 100 rows of sales data.")
print("Now run: python readexcelmain.py")
